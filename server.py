# -*- coding: utf-8 -*-
# pyre-unsafe
import sys
import os
import sqlite3
# 현재 디렉토리를 경로에 추가
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, HTTPException, UploadFile, File, Body, Request  # type: ignore
from fastapi.staticfiles import StaticFiles  # type: ignore
from fastapi.responses import StreamingResponse, JSONResponse  # type: ignore
from typing import List, Dict, Any, Optional, no_type_check
from pydantic import BaseModel
import json
import uvicorn  # type: ignore

# 커스텀 모듈 임포트
from database import get_db_conn  # type: ignore
from models import Record, StatusUpdate, Schedule  # type: ignore
import excel_service  # type: ignore

from fastapi.middleware.cors import CORSMiddleware  # type: ignore

def db_row_to_dict(row):
    """SQLite Row 또는 RealDictCursor Row를 딕셔너리로 변환"""
    if row is None:
        return None
    return dict(row)

app = FastAPI(title="지정폐기물 관리 시스템 API")

# 보안 미들웨어 비활성화 - Render 사이트에서 직접 데이터 관리 허용

# CORS 설정 추가 (브라우저 연결 안정성 확보)
origins = [
    "https://waste-management-ee09a.web.app",
    "https://waste-management-ee09a.firebaseapp.com",
    "https://waste-api-3j2l.onrender.com",
    "http://localhost:8000",
    "http://127.0.0.1:8000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 모든 응답에 CORS 헤더를 강제로 추가하는 미들웨어 (백업용)
@app.middleware("http")
async def add_cors_header(request: Request, call_next):
    response = await call_next(request)
    origin = request.headers.get("origin")
    if origin in origins:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
    return response

@app.get("/api/health")
def health_check():
    return {"status": "healthy", "environment": os.environ.get("RENDER", "local")}

@app.get("/api/debug-db")
def debug_db():
    try:
        conn = get_db_conn()
        res = {"type": str(type(conn))}
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) as count FROM records")
        res["records_count"] = cursor.fetchone()["count"]
        cursor.execute("SELECT COUNT(*) as count FROM schedules")
        res["schedules_count"] = cursor.fetchone()["count"]
        conn.close()
        return res
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/reset-db")
@app.post("/api/reset-db")
def reset_database():
    """DB 초기화: 모든 테이블의 데이터를 삭제하고 초기 데이터를 다시 로드합니다."""
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        tables = ["records", "schedules", "liquid_waste"]
        for table in tables:
            cursor.execute(f"DELETE FROM {table}")
        conn.commit()
        conn.close()
        
        # 데이터 삭제 후 즉시 자동 시딩 실행
        auto_seed_db()
        
        return {"message": "Database reset and seeded successfully", "tables_cleared": tables}
    except Exception as e:
        if conn:
            conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/records")
def get_records():
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM records ORDER BY date DESC, id DESC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

@app.post("/api/records")
def create_record(record: Record):
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        cursor.execute('''
        INSERT INTO records (slip_no, date, waste_type, amount, carrier, vehicle_no, processor, note1, note2, category, supplier, status, is_local)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
        RETURNING id
        ''', (record.slip_no, record.date, record.waste_type, record.amount, record.carrier, 
              record.vehicle_no, record.processor, record.note1, record.note2, record.category, record.supplier, record.status))
        conn.commit()
        new_id = cursor.fetchone()['id'] if os.environ.get('DATABASE_URL') else cursor.lastrowid
        conn.close()
        return {"message": "Success", "id": new_id}
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=400, detail="이미 존재하는 전표번호입니다. (중복 등록 불가)")
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/records/{record_id}")
def update_record(record_id: int, record: Record):
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        cursor.execute('''
        UPDATE records 
        SET slip_no = %s, date = %s, waste_type = %s, amount = %s, carrier = %s, 
            vehicle_no = %s, processor = %s, note1 = %s, note2 = %s, category = %s, 
            supplier = %s, status = %s
        WHERE id = %s
        ''', (record.slip_no, record.date, record.waste_type, record.amount, record.carrier, 
              record.vehicle_no, record.processor, record.note1, record.note2, record.category, 
              record.supplier, record.status, record_id))
        conn.commit()
        affected = cursor.rowcount
        conn.close()
        if affected == 0:
            raise HTTPException(status_code=404, detail="Record not found")
        return {"message": "Updated"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))

@app.patch("/api/records/{record_id}/status")
def update_status(record_id: int, status_update: StatusUpdate):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("UPDATE records SET status = %s WHERE id = %s", (status_update.status, record_id))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"message": "Updated"}

@app.delete("/api/records/{record_id}")
def delete_record(record_id: int):
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM records WHERE id = %s", (record_id,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"message": "Deleted"}

@app.delete("/api/records")
def delete_all_records():
    """모든 데이터 삭제 (관리자 전용)"""
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM records")
    conn.commit()
    conn.close()
    return {"message": "All records deleted"}


# --- 일정 관리 API ---

@app.get("/api/schedules")
def get_schedules():
    """모든 일정 조회"""
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM schedules ORDER BY date ASC, id ASC")
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]

@app.post("/api/schedules")
def create_schedule(schedule: Schedule):
    """새 일정 등록"""
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "INSERT INTO schedules (date, content, status) VALUES (%s, %s, %s) RETURNING id",
            (schedule.date, schedule.content, schedule.status or 'pending')
        )
        conn.commit()
        new_id = cursor.fetchone()['id'] if os.environ.get('DATABASE_URL') else cursor.lastrowid
        conn.close()
        return {"message": "Success", "id": new_id}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/schedules/{schedule_id}")
def update_schedule(schedule_id: int, schedule: Schedule):
    """일정 수정 (내용, 상태)"""
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE schedules SET date = %s, content = %s, status = %s WHERE id = %s",
            (schedule.date, schedule.content, schedule.status, schedule_id)
        )
        conn.commit()
        affected = cursor.rowcount
        conn.close()
        if affected == 0:
            raise HTTPException(status_code=404, detail="Schedule not found")
        return {"message": "Updated"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: int):
    """일정 삭제"""
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM schedules WHERE id = %s", (schedule_id,))
    conn.commit()
    affected = cursor.rowcount
    conn.close()
    if affected == 0:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return {"message": "Deleted"}

@app.get("/api/master")
def get_master():
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        
        cursor.execute("SELECT DISTINCT waste_type FROM records WHERE waste_type IS NOT NULL AND waste_type != ''")
        waste_types = [db_row_to_dict(row)['waste_type'] for row in cursor.fetchall()]
        
        cursor.execute("SELECT DISTINCT processor FROM records WHERE processor IS NOT NULL AND processor != ''")
        processors = [db_row_to_dict(row)['processor'] for row in cursor.fetchall()]
        
        cursor.execute("SELECT DISTINCT vehicle_no FROM records WHERE vehicle_no IS NOT NULL AND vehicle_no != ''")
        vehicles = [db_row_to_dict(row)['vehicle_no'] for row in cursor.fetchall()]

        conn.close()
        return {
            "wasteTypes": sorted(waste_types),
            "processors": sorted(processors),
            "vehicles": sorted(vehicles)
        }
    except Exception as e:
        import traceback
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": traceback.format_exc()})

# --- 데이터 입출력 (Import/Export) API ---

@app.get("/api/export/excel")
def export_excel():
    """DB 데이터를 엑셀로 추출"""
    try:
        output = excel_service.export_to_excel()
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=waste_records_export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export/excel/filtered")
async def export_excel_filtered(data: List[Dict[str, Any]] = Body(...)):
    """필터링된 데이터를 엑셀로 추출"""
    try:
        output = excel_service.export_filtered_to_excel(data)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=waste_records_filtered.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """엑셀 파일 업로드 및 데이터 저장"""
    try:
        content = await file.read()
        result = excel_service.import_from_excel(content)
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 처리 중 오류: {str(e)}")

@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...)):
    """CSV 파일 업로드 및 데이터 저장"""
    try:
        content = await file.read()
        result = excel_service.import_from_csv(content)
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV 처리 중 오류: {str(e)}")

# --- 팀별 액상폐기물 관리 API ---

@app.post("/api/liquid-waste/upload")
@no_type_check
async def upload_liquid_waste(file: UploadFile = File(...)):
    """액상폐기물 Excel 파일 업로드 및 파싱"""
    try:
        import openpyxl  # type: ignore
        from io import BytesIO
        
        content = await file.read()
        # Any 타입을 사용하여 Pyre2의 타입 추론 강제 중단
        wb: Any = openpyxl.load_workbook(BytesIO(content), data_only=True)  # type: ignore
        
        if wb is None:
            raise HTTPException(status_code=400, detail="워크북을 로드할 수 없습니다.")
            
        # 입고리스트 시트 찾기 (시트명에 '입고리스트' 포함)
        target_sheet = None
        year_month = None
        for sname in wb.sheetnames:  # type: ignore
            if sname and '입고리스트' in str(sname):
                # 시트명에서 연월 추출 (예: "26.1 팀별 액상폐기물 입고리스트" → "2026-01")
                parts = str(sname).split()
                if parts:
                    ym = parts[0]  # "26.1"
                    ym_parts = ym.split('.')
                    if len(ym_parts) == 2:
                        try:
                            y = int(ym_parts[0])
                            m = int(ym_parts[1])
                            year = y + 2000 if y < 100 else y
                            year_month = f"{year}-{m:02d}"
                            target_sheet = wb[sname]  # type: ignore
                            break
                        except ValueError:
                            continue
        
        if target_sheet is None or year_month is None:
            raise HTTPException(status_code=400, detail="입고리스트 시트를 찾을 수 없습니다.")
        
        # Any 타입을 사용하여 Pyre2의 속성 접근 오류 차단
        ws: Any = target_sheet
        
        # 헤더 행 찾기 (배출일이 있는 행)
        header_row: Optional[int] = None
        # max_row가 None일 경우 대비 (타입 안정성 확보)
        max_r: int = 100
        max_c: int = 20
        if ws:
            try:
                if hasattr(ws, 'max_row') and ws.max_row:
                    max_r = int(ws.max_row)
                if hasattr(ws, 'max_column') and ws.max_column:
                    max_c = int(ws.max_column)
            except:
                pass
        
        for r in range(1, min(max_r + 1, 10)):  # type: ignore
            for c in range(1, min(max_c + 1, 10)):  # type: ignore
                v = ws.cell(row=r, column=c).value  # type: ignore
                if v and '배출일' in str(v):
                    header_row = int(r)
                    break
            if header_row:
                break
        
        if not header_row:
            raise HTTPException(status_code=400, detail="헤더 행(배출일)을 찾을 수 없습니다.")
        
        # 데이터 행 파싱 (헤더 다음 행부터, 빈 행 건너뛰기)
        records = []
        # 재고 행 건너뛰기 (데이터 시작 전 재고 행이 있을 수 있음)
        # Type Checker가 header_row가 int임을 보장하도록 체크
        if header_row is None:
             raise HTTPException(status_code=400, detail="헤더 행을 찾을 수 없습니다.")
             
        # 강제 형변환으로 산술 연산 오류 제거
        h_row: int = int(header_row)
        data_start: int = h_row + 1
        for r in range(data_start, max_r + 1):  # type: ignore
            # 최소한 배출부서(E열)과 반입량(H열)이 있어야 유효 데이터
            team_cell = ws.cell(row=r, column=5).value  # type: ignore
            amount_cell = ws.cell(row=r, column=8).value  # type: ignore
            
            # None 체크
            team = str(team_cell).strip() if team_cell else None
            amount = amount_cell
            
            if not team or not amount:
                # 재고 행이거나 빈 행이면 건너뛰기
                cell_h_val = ws.cell(row=r, column=8).value  # type: ignore
                if cell_h_val and '재고' in str(cell_h_val):
                    continue
                if not team:
                    continue
            
            discharge_date = ws.cell(row=r, column=1).value  # type: ignore
            receive_date = ws.cell(row=r, column=2).value  # type: ignore
            waste_type = ws.cell(row=r, column=3).value  # type: ignore
            content_val = ws.cell(row=r, column=4).value  # type: ignore
            discharger = ws.cell(row=r, column=6).value  # type: ignore
            quantity = ws.cell(row=r, column=7).value  # type: ignore
            amount_k = ws.cell(row=r, column=8).value  # type: ignore
            
            # 날짜 변환
            if hasattr(discharge_date, 'strftime'):
                discharge_date = discharge_date.strftime('%Y-%m-%d')
            else:
                discharge_date = str(discharge_date) if discharge_date else None
            
            if hasattr(receive_date, 'strftime'):
                receive_date = receive_date.strftime('%Y-%m-%d')
            else:
                receive_date = str(receive_date) if receive_date else None
            
            amount_kg = float(amount) if amount else 0  # type: ignore
            quantity_ea = int(quantity) if quantity else 0  # type: ignore
            
            records.append({  # type: ignore
                'year_month': str(year_month),
                'discharge_date': discharge_date,
                'receive_date': receive_date,
                'waste_type': str(waste_type) if waste_type else '',
                'content': str(content_val) if content_val else '',
                'team': str(team).strip(),
                'discharger': str(discharger) if discharger else '',
                'quantity_ea': quantity_ea,
                'amount_kg': amount_kg
            })
        
        if not records:
            raise HTTPException(status_code=400, detail="파싱된 데이터가 없습니다.")
        
        # DB 저장 (기존 해당 월 데이터 삭제 후 삽입)
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM liquid_waste WHERE year_month = %s", (year_month,))
        
        for rec in records:
            cursor.execute("""
                INSERT INTO liquid_waste 
                (year_month, discharge_date, receive_date, waste_type, content, team, discharger, quantity_ea, amount_kg)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (rec['year_month'], rec['discharge_date'], rec['receive_date'],
                  rec['waste_type'], rec['content'], rec['team'], rec['discharger'],
                  rec['quantity_ea'], rec['amount_kg']))
        
        conn.commit()
        conn.close()
        
        return {
            "success": True,
            "year_month": year_month,
            "count": len(records),
            "message": f"{year_month} 데이터 {len(records)}건 저장 완료"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 처리 중 오류: {str(e)}")

@app.get("/api/liquid-waste")
async def get_liquid_waste(year: str | None = None):
    """액상폐기물 데이터 조회"""
    conn = get_db_conn()
    cursor = conn.cursor()
    
    if year:
        cursor.execute("""
            SELECT * FROM liquid_waste 
            WHERE year_month LIKE %s 
            ORDER BY year_month, team, receive_date
        """, (f"{year}-%",))
    else:
        cursor.execute("SELECT * FROM liquid_waste ORDER BY year_month, team, receive_date")
    
    rows = cursor.fetchall()
    conn.close()
    
    return [dict(row) for row in rows]


@app.delete("/api/liquid-waste/{year_month}")
async def delete_liquid_waste(year_month: str):
    """특정 월 액상폐기물 데이터 삭제"""
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM liquid_waste WHERE year_month = %s", (year_month,))
    deleted = cursor.rowcount
    conn.commit()
    conn.close()
    
    if deleted == 0:
        raise HTTPException(status_code=404, detail=f"{year_month} 데이터가 없습니다.")
    
    return {"success": True, "deleted": deleted, "message": f"{year_month} 데이터 {deleted}건 삭제"}

# --- 마이그레이션 전용 API ---

class LiquidWasteRecord(BaseModel):
    year_month: str
    discharge_date: Optional[str] = None
    receive_date: Optional[str] = None
    waste_type: str = ""
    content: str = ""
    team: str
    discharger: str = ""
    quantity_ea: int = 0
    amount_kg: float = 0.0

@app.post("/api/liquid-waste/migration")
def migrate_liquid_waste(records: List[LiquidWasteRecord]):
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        count = 0
        for r in records:
            cursor.execute("""
                INSERT INTO liquid_waste 
                (year_month, discharge_date, receive_date, waste_type, content, team, discharger, quantity_ea, amount_kg)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (r.year_month, r.discharge_date, r.receive_date,
                  r.waste_type, r.content, r.team, r.discharger,
                  r.quantity_ea, r.amount_kg))
            count += 1
        conn.commit()
        conn.close()
        return {"message": f"Migrated {count} records"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/schedules/migration")
def migrate_schedules(schedules: List[Schedule]):
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        count = 0
        for s in schedules:
            cursor.execute(
                "INSERT INTO schedules (date, content, status) VALUES (%s, %s, %s)",
                (s.date, s.content, s.status or 'pending')
            )
            count += 1
        conn.commit()
        conn.close()
        return {"message": f"Migrated {count} schedules"}
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

def auto_seed_db():
    """DB가 비어있을 경우 로컬 JSON 파일들로부터 데이터를 자동으로 채움"""
    conn = get_db_conn()
    cursor = conn.cursor()
    
    try:
        # 1. Records (인계서)
        # 1. Records (인계서)
        # 1. Records (인계서)
        cursor.execute("SELECT COUNT(*) as count FROM records")
        first_row = cursor.fetchone()
        records_count = first_row['count'] if isinstance(first_row, dict) else first_row[0]
        
        # 데이터가 없거나, 처리량이 0이거나, 날짜가 비어있는 데이터가 있다면 재동기화 수행
        cursor.execute("SELECT COUNT(*) as count FROM records WHERE amount = 0 OR date = '' OR date IS NULL")
        fix_needed_row = cursor.fetchone()
        fix_needed_count = fix_needed_row['count'] if isinstance(fix_needed_row, dict) else fix_needed_row[0]

        if records_count == 0 or fix_needed_count > 0:
            file_path = "render_records.json"
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    
                    # 로컬 데이터 매핑 및 is_local=1 강제 설정 (배포환경에서도 수정 가능하도록)
                    for r in data:
                        # amount, is_local 값을 명시적으로 처리
                        amount = float(r.get('amount', 0) or 0)
                        
                        cursor.execute("""
                            INSERT INTO records (slip_no, date, waste_type, amount, carrier, vehicle_no, processor, note1, note2, category, supplier, status, is_local, created_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 1, %s)
                            ON CONFLICT (slip_no) DO UPDATE SET
                                date = EXCLUDED.date,
                                amount = EXCLUDED.amount,
                                is_local = 1,
                                waste_type = EXCLUDED.waste_type,
                                processor = EXCLUDED.processor,
                                vehicle_no = EXCLUDED.vehicle_no
                        """, (
                            r.get('slip_no', ''), 
                            r.get('date', ''), 
                            r.get('waste_type', ''), 
                            amount, 
                            r.get('carrier', ''), 
                            r.get('vehicle_no', ''), 
                            r.get('processor', ''), 
                            r.get('note1', ''), 
                            r.get('note2', ''), 
                            r.get('category', ''), 
                            r.get('supplier', ''), 
                            r.get('status', 'completed'), 
                            r.get('created_at', '')
                        ))
                print(f"✅ Records seeded: {len(data)} items (with is_local=1)")

        # 2. Schedules (일정)
        # 2. Schedules (일정)
        cursor.execute("SELECT COUNT(*) as count FROM schedules")
        first_row = cursor.fetchone()
        schedules_count = first_row['count'] if isinstance(first_row, dict) else first_row[0]
        
        if schedules_count == 0:
            file_path = "local_schedules.json"
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    for s in data:
                        cursor.execute(
                            "INSERT INTO schedules (date, content, status, created_at) VALUES (%s, %s, %s, %s)",
                            (s.get('date', ''), s.get('content', ''), s.get('status', 'pending'), s.get('created_at', ''))
                        )
                print(f"✅ Schedules seeded: {len(data)} items")

        # 3. Liquid Waste (액상폐기물 - local_liquid_waste.json)
        cursor.execute("SELECT COUNT(*) as count FROM liquid_waste")
        first_row = cursor.fetchone()
        lw_count = first_row['count'] if isinstance(first_row, dict) else first_row[0]
        
        if lw_count == 0:
            file_path = "local_liquid_waste.json"
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    for lw in data:
                        cursor.execute("""
                            INSERT INTO liquid_waste 
                            (year_month, discharge_date, receive_date, waste_type, content, team, discharger, quantity_ea, amount_kg)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            lw.get('year_month', ''),
                            lw.get('discharge_date', None),
                            lw.get('receive_date', None),
                            lw.get('waste_type', ''),
                            lw.get('content', ''),
                            lw.get('team', ''),
                            lw.get('discharger', ''),
                            lw.get('quantity_ea', 0),
                            float(lw.get('amount_kg', 0) or 0)
                        ))
                print(f"✅ Liquid waste seeded: {len(data)} items")

        conn.commit()
    except Exception as e:
        print(f"❌ Auto-seeding failed: {e}")
    finally:
        conn.close()

@app.on_event("startup")
async def startup_event():
    # Render 환경에서만 실행하거나 로컬에서도 실행 가능
    # Auto-seeding logic triggered
    auto_seed_db()

app.mount("/", StaticFiles(directory="frontend", html=True), name="static")

if __name__ == "__main__":
    import uvicorn  # type: ignore
    # Render(클라우드) 환경에서는 PORT 환경변수를 사용해야 하며, 로컬에서는 8000을 기본값으로 사용합니다.
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port, access_log=True)  # type: ignore
