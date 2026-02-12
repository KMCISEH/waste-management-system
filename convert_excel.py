# -*- coding: utf-8 -*-
"""
지정폐기물 전자인계서 엑셀 → SQLite 데이터베이스 마이그레이션 스크립트
실행: python convert_excel.py
"""
import sqlite3
import os
from datetime import datetime
import openpyxl

EXCEL_PATH = r"P:\안전환경팀\3. 환경PART\3. 폐기물관리법\1. 폐기물 년도별 처리실적(발생시 당일입력)\지정폐기물 전자인계서(2017년 이후~).xlsx"
DB_PATH = "waste_management.db"

def migrate():
    if not os.path.exists(DB_PATH):
        print("❌ 데이터베이스 파일이 없습니다. init_db.py를 먼저 실행하세요.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.worksheets[0]  # 2026년 시트

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 기존 데이터 수 확인 (전표번호 기준 중복 체크를 위함)
    cursor.execute("SELECT slip_no FROM records")
    existing_slips = {row[0] for row in cursor.fetchall()}

    records_added = 0
    records_skipped = 0

    # 헤더 제외 키워드
    skip_values = {"전표 NO.", "폐기물명", "처리 업체", "폐기물종류", "처리자명", "차량 번호", "인계번호"}

    for r in range(3, ws.max_row + 1):
        slip_no = ws.cell(row=r, column=2).value
        if slip_no is None:
            break

        slip_str = str(slip_no).strip()
        if slip_str in existing_slips or not slip_str:
            records_skipped += 1
            continue
        
        waste_name = ws.cell(row=r, column=4).value
        # 중간 헤더 행 건너뛰기
        if waste_name in skip_values:
            continue

        date_val = ws.cell(row=r, column=3).value
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val).replace(".", "-")
        else:
            date_str = ""

        amount = ws.cell(row=r, column=5).value or 0
        carrier = ws.cell(row=r, column=6).value or ""
        vehicle = ws.cell(row=r, column=7).value or ""
        processor = ws.cell(row=r, column=8).value or ""
        note1 = ws.cell(row=r, column=9).value or ""
        note2 = ws.cell(row=r, column=10).value or ""
        category = ws.cell(row=r, column=11).value or ""
        supplier = ws.cell(row=r, column=12).value or ""

        try:
            cursor.execute('''
            INSERT INTO records (slip_no, date, waste_type, amount, carrier, vehicle_no, processor, note1, note2, category, supplier, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'completed')
            ''', (slip_str, date_str, waste_name, amount, str(carrier), str(vehicle), str(processor), str(note1), str(note2), str(category), str(supplier)))
            records_added += 1
            existing_slips.add(slip_str)
        except sqlite3.IntegrityError:
            records_skipped += 1

    conn.commit()
    conn.close()

    print(f"✅ 마이그레이션 완료!")
    print(f"   - 신규 추가: {records_added}건")
    print(f"   - 중복 제외: {records_skipped}건")
    print(f"   - 데이터베이스: {DB_PATH}")

if __name__ == "__main__":
    migrate()
