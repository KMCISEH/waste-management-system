# -*- coding: utf-8 -*-
import pandas as pd
import sqlite3
import os
from io import BytesIO
from database import get_db_conn

def export_to_excel():
    """DB의 모든 데이터를 한글 컬럼명이 포함된 엑셀 파일로 변환"""
    conn = get_db_conn()
    # PostgreSQL/SQLite 모두 호환되도록 AS 별칭에 큰따옴표 사용
    query = """
        SELECT 
            date AS "처리일",
            slip_no AS "전표번호",
            waste_type AS "폐기물명",
            amount AS "처리량(톤)",
            vehicle_no AS "차량번호",
            processor AS "처리업체",
            note1 AS "처리방법",
            category AS "비고",
            supplier AS "장소"
        FROM records 
        ORDER BY date DESC, id DESC
    """
    # pandas.read_sql_query는 실제 DB 연결 객체가 필요 (래퍼가 아닌 원본)
    raw_conn = conn._conn if hasattr(conn, '_conn') else conn
    df = pd.read_sql_query(query, raw_conn)
    conn.close()
    
    # 기본값 설정
    df['장소'] = df['장소'].fillna('공장')
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='AllRecords')
    
    output.seek(0)
    return output

def export_filtered_to_excel(data):
    """프런트엔드에서 필터링된 데이터를 전달받아 엑셀 생성"""
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='FilteredRecords')
    
    output.seek(0)
    return output


def export_cost_settlement_excel(data: dict):
    """비용 정산 엑셀 파일 생성 (스타일링 포함)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "비용정산"

    # 스타일 정의
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name='맑은 고딕', size=18, bold=True, color='222222')
    section_font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
    header_font = Font(name='맑은 고딕', size=10, bold=True, color='444444')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    data_font = Font(name='맑은 고딕', size=10, color='333333')
    amount_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    subtotal_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    subtotal_font = Font(name='맑은 고딕', size=11, bold=True, color='276A3C')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    num_fmt = '#,##0'
    ton_fmt = '#,##0.00'

    year_month = data.get('yearMonth', '')
    ym_parts = year_month.split('-') if year_month else ['', '']
    title_text = f"{ym_parts[0]}년 {int(ym_parts[1]) if len(ym_parts) > 1 and ym_parts[1] else ''}월 지정폐기물 처리비 및 잡이익"

    # 열 너비 설정
    col_widths = [16, 28, 8, 12, 14, 16, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def apply_border_row(row_num, cols=8):
        for c in range(1, cols + 1):
            ws.cell(row=row_num, column=c).border = thin_border

    def write_title(row):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=title_text)
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 36
        return row + 1

    def write_section_header(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 28
        apply_border_row(row)
        return row + 1

    def write_table_header(row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[row].height = 30
        return row + 1

    def write_data_row(row, values, amount_col=None, is_subtotal=False, amount_fmt=None):
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = subtotal_font if is_subtotal else data_font
            cell.border = thin_border
            if is_subtotal:
                cell.fill = subtotal_fill
            if isinstance(v, (int, float)):
                if c == amount_col:
                    cell.number_format = amount_fmt if amount_fmt else ton_fmt
                    if not is_subtotal and v != 0:
                        cell.fill = amount_fill
                    cell.alignment = right_align
                elif isinstance(v, float):
                    cell.number_format = ton_fmt
                    cell.alignment = right_align
                else:
                    cell.number_format = num_fmt
                    cell.alignment = right_align
            elif c == 1 and is_subtotal:
                cell.alignment = center
            else:
                cell.alignment = left_align if c <= 2 else center
        
        if is_subtotal:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            
        ws.row_dimensions[row].height = 22
        return row + 1

    # === 1. 타이틀 ===
    row = write_title(1)

    # === 2. 폐기물 처리비용 섹션 ===
    row = write_section_header(row, '폐기물 처리비용')
    row = write_table_header(row, ['업체명', '폐기물명', '성상', '분류 번호',
                                     '처리량\n(TON)', '단가(TON/원)', '처리비용\n(공급가기준)', '비 고'])

    process_items = data.get('processing', [])
    process_total = 0
    process_amount_total = 0
    for item in process_items:
        cost = item.get('cost', 0)
        amount = item.get('amount', 0)
        process_total += cost
        process_amount_total += amount
        row = write_data_row(row, [
            item.get('processor', ''),
            item.get('wasteName', ''),
            item.get('phase', ''),
            item.get('classCode', ''),
            amount,
            item.get('unitCost', 0),
            cost,
            item.get('note', ''),
        ], amount_col=5)

    # 처리비용 소계
    row = write_data_row(row, [
        '처리비용 소계', '', '', '', process_amount_total, '', process_total, ''
    ], is_subtotal=True, amount_col=5)

    row += 1  # 빈 행

    # === 3. 폐기물 운반별도비용 섹션 ===
    row = write_section_header(row, '폐기물 운반별도비용')
    row = write_table_header(row, ['업체명', '폐기물명', '성상', '분류 번호',
                                     '처리량\n(TON)', '단가(TON/원)', '운반비용\n(공급가기준)', '비 고'])

    transport_items = data.get('transport', [])
    transport_total = 0
    transport_amount_total = 0
    for item in transport_items:
        cost = item.get('cost', 0)
        amount = item.get('amount', 0)
        transport_total += cost
        transport_amount_total += amount
        row = write_data_row(row, [
            item.get('carrier', ''),
            item.get('wasteName', ''),
            item.get('phase', ''),
            item.get('classCode', ''),
            amount,
            item.get('unitCost', 0),
            cost,
            item.get('note', ''),
        ], amount_col=5)

    # 운반비용 소계
    row = write_data_row(row, [
        '운반별도비용 소계', '', '', '', transport_amount_total, '', transport_total, ''
    ], is_subtotal=True, amount_col=5)

    row += 1  # 빈 행

    # === 4. 폐기물 잡이익비용 섹션 ===
    row = write_section_header(row, '폐기물 잡이익비용')
    row = write_table_header(row, ['업체명', '폐기물명', '성상', '분류 번호',
                                     'EA', '단가(EA/원)', '매각비용\n(공급가기준)', '비 고'])

    revenue_items = data.get('revenue', [])
    revenue_total = 0
    revenue_ea_total = 0
    for item in revenue_items:
        cost = item.get('cost', 0)
        ea = item.get('ea', 0)
        revenue_total += cost
        revenue_ea_total += ea
        row = write_data_row(row, [
            item.get('processor', ''),
            item.get('wasteName', ''),
            item.get('type', ''),
            item.get('classCode', ''),
            ea,
            item.get('unitCost', 0),
            cost,
            item.get('note', ''),
        ], amount_col=5, amount_fmt=num_fmt)

    # 잡이익 소계
    row = write_data_row(row, [
        '잡이익비용 소계', '', '', '', revenue_ea_total, '', revenue_total, ''
    ], is_subtotal=True, amount_col=5, amount_fmt=num_fmt)

    # 인쇄 설정
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_from_excel(file_content: bytes):
    """엑셀 파일에서 데이터를 읽어 DB에 저장"""
    read_errors = []
    file_stream = BytesIO(file_content)
    
    # 스타일 메타데이터 이슈(styleId) 회피를 위해 calamine 우선 시도
    for engine in ("calamine", "openpyxl"):
        try:
            file_stream.seek(0)
            df = pd.read_excel(file_stream, engine=engine)
            return _import_dataframe(df)
        except Exception as e:
            read_errors.append(f"{engine}: {str(e)}")

    raise ValueError(f"엑셀 처리 중 오류: {' | '.join(read_errors)}")

def import_from_csv(file_content: bytes):
    """CSV 파일에서 데이터를 읽어 DB에 저장"""
    # 한글 인코딩 고려 (cp949 또는 utf-8)
    try:
        df = pd.read_csv(BytesIO(file_content), encoding='utf-8')
    except:
        df = pd.read_csv(BytesIO(file_content), encoding='cp949')
    return _import_dataframe(df)

def _import_dataframe(df):
    """DataFrame 데이터를 DB에 병합 (중복 제외)"""
    conn = get_db_conn()
    cursor = conn.cursor()
    is_postgres = bool(os.environ.get('DATABASE_URL'))
    
    # === 컬럼명 정규화: 앞뒤 공백 및 줄바꿈 제거 ===
    df.columns = [str(c).strip().replace('\n', ' ') for c in df.columns]
    
    # NaN 처리 및 유틸리티 함수 정의
    def get_val(r, keys, default=''):
        for k in keys:
            v = r.get(k)
            if v is not None and pd.notnull(v):
                return v
        return default

    added_count = 0
    skipped_count = 0
    error_samples = []
    
    for idx, row in df.iterrows():
        try:
            # 기본값 설정 및 타입 변환
            # 전표번호 매칭: 다양한 컬럼명 지원
            slip_no = str(get_val(row, [
                'slip_no', '전표번호', '인계번호', '인계번호(*)', 
                '전자인계번호', '관리번호', 'No', 'no'
            ])).strip()
            if not slip_no or slip_no.lower() == 'nan' or slip_no == 'None':
                continue
                
            processor = str(get_val(row, [
                'processor', '처리업체', '처리자명', '처리업체명', '처리자명(*)'
            ])).strip()
            
            # 처리업체별 자동 분류(category) 매핑 로직
            category = str(get_val(row, ['category', '분류', '폐기물분류', '비고']))
            if not category or category.lower() == 'nan' or category == '':
                if '해동이앤티' in processor: category = "AO-Tar"
                elif '제일자원' in processor: category = "AO-TAR"
                elif '디에너지' in processor: category = "메탄올"
            
            values = (
                slip_no,
                str(get_val(row, ['date', '날짜', '처리일', '인계일자', '인계일자(*)', '일자'])),
                str(get_val(row, ['waste_type', '폐기물종류', '폐기물명', '폐기물종류(성상)', '폐기물종류(성상)(*)', '품명'])),
                float(get_val(row, ['amount', '중량', '처리량', '처리량(톤)', '위탁량', '위탁량(*)', '수량'], 0)),
                str(get_val(row, ['carrier', '운반업체', '운반자명', '운반업체명', '운반자명(*)'])),
                str(get_val(row, ['vehicle_no', '차량번호', '차량 번호'])),
                processor,
                str(get_val(row, ['note1', '비고1', '처리방법', '비고'])),
                str(get_val(row, ['note2', '비고2'])),
                category,
                str(get_val(row, ['supplier', '공급업체', '장소'])),
                str(get_val(row, ['status'], 'completed'))
            )
            
            if is_postgres:
                # PostgreSQL: SAVEPOINT로 중복 에러가 트랜잭션을 깨지 않도록 보호
                cursor.execute("SAVEPOINT import_row")
                try:
                    cursor.execute('''
                    INSERT INTO records (slip_no, date, waste_type, amount, carrier, vehicle_no, processor, note1, note2, category, supplier, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', values)
                    cursor.execute("RELEASE SAVEPOINT import_row")
                    added_count += 1
                except Exception:
                    cursor.execute("ROLLBACK TO SAVEPOINT import_row")
                    skipped_count += 1
            else:
                # SQLite: 개별 INSERT 에러가 트랜잭션에 영향 없음
                cursor.execute('''
                INSERT INTO records (slip_no, date, waste_type, amount, carrier, vehicle_no, processor, note1, note2, category, supplier, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', values)
                added_count += 1
                
        except Exception as e:
            skipped_count += 1
            if len(error_samples) < 3:
                error_samples.append(f"Row {idx}: {str(e)}")
            
    conn.commit()
    conn.close()
    
    result = {"added": added_count, "skipped": skipped_count, "columns": list(df.columns)}
    if error_samples:
        result["error_samples"] = error_samples
    return result
